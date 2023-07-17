from openpyxl import Workbook, load_workbook

gene_list1 = ['GFAP', 'ATRX', 'IDH1/2', 'IDH1/2', 'TP53', '(promotor)', '1p/19q', 'chr7', 'chr10', 'CDKN2A/B', 'EGFR', 'ki67', 'ATRX',	'CD5', 'CD20', 'met_MGMT', 'CD68'	, 'Olig2','EMA', 'BRAF', 'CD34']
gene_list2 = ['GFAP', 'ATRX', 'IDH1/2', 'IDH1/2', 'TP53', 'TERT', '1p/19q', 'chr7', 'chr10', 'CDKN2A/B', 'EGFR', 'ki67', 'ATRX',	'CD5', 'CD20', 'met_MGMT', 'CD68'	, 'Olig2','EMA', 'BRAF', 'CD34']
gene_list3 = ['GFAP', 'ATRX', 'IDH1/2', 'IDH1/2', 'TP53', 'TERT', '1p/19q', '7', '10', 'CDKN2A/B', 'EGFR', 'ki67', 'ATRX',	'CD5', 'CD20', 'met_MGMT', 'CD68'	, 'Olig2','EMA', 'BRAF', 'CD34']

# Load the workbook
workbook = load_workbook('Glioma-XAI-kopia.xlsx')

# Select the active sheet
sheet = workbook.active

def column_number_to_name(column_number):
    dividend = column_number
    column_name = ''

    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_name = chr(65 + modulo) + column_name
        dividend = (dividend - modulo) // 26

    return column_name

def rodzaj(words, current_row, gene_column, gene):
  for index, word in enumerate(words):
    if (word == gene):
      if index + 1 < len(words):
        if words[index + 1] == 'wykryto':
          sheet[column_number_to_name(gene_column) + str(current_row)] = 1
          break
        elif words[index + 1] == 'brak':
          sheet[column_number_to_name(gene_column) + str(current_row)] = 0
          break

def uwagi(words, current_row, gene_column, gene, wykryto_index, niewykryto_index):
  if niewykryto_index == 0:
    for word in words:
      if (word == gene):
        sheet[column_number_to_name(gene_column) + str(current_row)] = 1
        break
  else:
    for i in range(wykryto_index, niewykryto_index):
      if (words[i] == gene):
        sheet[column_number_to_name(gene_column) + str(current_row)] = 1
        break
    for i in range(niewykryto_index, len(words)):
      if (words[i] == gene):
        sheet[column_number_to_name(gene_column) + str(current_row)] = 0
        break


for i, row in enumerate(sheet.iter_rows(min_row=44, min_col=19, max_col=19, values_only=True)):
    description = str(row[0])
    description = description.replace('.', ' ').replace(',', ' ')
    words = description.split()
    current_row = i + 44
    current_column = 19

    if words[0] == 'Rodzaj':
      for index, gene in enumerate(gene_list1):
        rodzaj(words, current_row, current_column + index + 1, gene)
      for index, gene in enumerate(gene_list2):
        rodzaj(words, current_row, current_column + index + 1, gene)
    
    if words[0] == 'Uwagi:':
      wykryto_index = 0
      niewykryto_index = 0
      for index, word in enumerate(words):
        if word == 'wykryto:' or word == 'wykryto':
          wykryto_index = index
          break
      for index, word in enumerate(words):
        if word == 'Nie':
          niewykryto_index = index
          break
      for index, gene in enumerate(gene_list3):
        uwagi(words, current_row, current_column + index + 1, gene, wykryto_index, niewykryto_index)

for row in sheet.iter_rows(min_row=44, max_row=131, min_col=20, max_col=40):
  for cell in row:
    if cell.value is None:
      cell.value = '-'

workbook.save('Glioma-XAI-kopia.xlsx')


#todo status metylacji MGMT
#todo ki67
